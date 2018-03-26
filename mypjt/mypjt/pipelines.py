# -*- coding: utf-8 -*-

import scrapy
import logging
import os
import sys
import openpyxl
from openpyxl import load_workbook,workbook,Workbook
from mypjt.items  import MypjtItem
import traceback
import pymysql
from  mypjt.Public_Module import init_mysql,currency_translate_dict,top_dict

logger=logging.getLogger(sys._getframe().f_code.co_filename)


#判断sheet是否存在表头，如果不存在就创建
def sheet_is_exsit_top(ws,top_list):
    if len(tuple(ws.rows))<=0:
        ws.append(top_list)
#在工作表中查找sheet,如果没找到就创建他 create_index<0,添加到末尾
def find_wb_sheet(wb,sheet_name,create_index):
    sheet_name_list=wb.get_sheet_names()
    res=compareListEle(sheet_name_list,sheet_name)
    if not res:
       if create_index<0:
          ws = wb.create_sheet(sheet_name)
       elif create_index>0:
          ws = wb.create_sheet(sheet_name, create_index)
       else:
          ws = wb.active
          ws.title=sheet_name
    else:
        ws=wb[sheet_name]
    return ws    

def compareListEle(list,item):
    if len(list)<=0:
        return False
    for ele in list:
        if ele==item:
           return True
    return False
def find_file(path,filename):
    for root,dirs,filesname in os.walk(path):
        for file in filesname:
            if file==filename:
                return True
    return  False


class MypjtPipeline(object):
    
    def __init__(self):
        #self.no_Update_datetime=True
        self.Data_wb=None
        self.excelFilePath=None
        self.conn=None
        self.cursor=None
        
    def init_excel_table(self,excel_name):
        logger.info('init excel')
        current_abspath=os.path.abspath('.')
        dataFilePath=os.path.join(current_abspath,'dataFiles')
        if not os.path.exists(dataFilePath):
           logger.info ("directory %s no exists， create new directory" % dataFilePath)
           os.mkdir(dataFilePath)
        self.excelFilePath=os.path.join(dataFilePath,excel_name+".xlsx")
        res=find_file(dataFilePath,excel_name+".xlsx")
        if res:
           self.Data_wb=load_workbook(self.excelFilePath)
        else:
           self.Data_wb=Workbook()
           #self.Data_wb.remove_sheet(self.Data_wb.get_sheet_by_name('Sheet1'))
    def update_datatime(self,datetime,CurrencyName):
        try :
            sql="update all_currency_tb set update_datetime='{0}' where currency_name='{1}' ".format(datetime,CurrencyName)
            self.cursor.execute(sql)
            self.conn.commit()
        except Exception as e:
            self.conn.rollback()
            logger.error('update {0} update_time failure,\n error massge: {1}'.format(CurrencyName,traceback.format_exc()))
        else:
            logger.info('update {0} update_time sucessful'.format(CurrencyName))
            self.no_Update_datetime=False

    def save_to_excel(self,item):
        if  self.Data_wb==None:
            self.init_excel_table(item['currency_name'])
        logger.info('write excel')
        top_list2=item['top_list'][1:len(item['top_list'])-1]
        allData_ws=find_wb_sheet(self.Data_wb,'allData_sheet',0)
        sheet_is_exsit_top(allData_ws,item['top_list'][:len(item['top_list'])-1])
        size=len(item['data_list'][0])
        for i in item['data_list'][::-1]:
            for line in i:
                allData_ws.append(line)
        for index in range(0,size):
            ws=find_wb_sheet(self.Data_wb,item['data_list'][0][index][0],-1)
            sheet_is_exsit_top(ws,top_list2)
            for i in item['data_list'][::-1]:
                line=i[index]
                ws.append(line[1:len(line)])
        logger.info('write excel finish')

        if  self.Data_wb!=None:
            self.Data_wb.save(self.excelFilePath)
        else:
            logger.error('save excel fuail,self.Data_wb is None ')
        logger.info('close excel')#日志
    def save_linedata_to_mysql(self,cur_tb_name,childtb_name,ex_cur_name,sql_list):
        try :
            effect_row=self.cursor.execute(sql_list[0])
            self.conn.commit()
        except Exception as e:
            self.conn.rollback()
            logger.error('create child table {0} failure,\n error massge: {1}'.format(childtb_name,traceback.format_exc()) )
        else:
            try :
                effect_row=self.cursor.execute(sql_list[1])
                self.conn.commit()
            except Exception as e:
                self.conn.rollback()
                logger.error('insert {0} to {1} failure,\n error massge: {2}'.format(ex_cur_name, childtb_name,traceback.format_exc()))
            else:
                effect_row=self.cursor.execute(sql_list[2])
                if  effect_row<=0:
                    try :
                        effect_row=self.cursor.execute(sql_list[3])
                        self.conn.commit()
                    except Exception as e:
                        self.conn.rollback()
                        logger.error('in {0} , insert {1} and {2}  record failure,\n error massge: {3}'.format(cur_tb_name,ex_cur_name,childtb_name,traceback.format_exc()))  
                    else:
                        logger.info('in {0} , insert {1} and {2}  record sucessful'.format(cur_tb_name,ex_cur_name,childtb_name))   
                else:
                    logger.info('in {0} ,{1} and {2}  record exists'.format(cur_tb_name,ex_cur_name,childtb_name))
    def insert_data_pre(self,sql_list,childtb_name):
        try :
            #print('sql_0 :'+sql_list[0])
            effect_row=self.cursor.execute(sql_list[0])
            self.conn.commit()
        except Exception as e:
            self.conn.rollback()
            logger.info("error sql massage :"+sql_list[0])
            logger.error('create child table {0} failure,\n error massge: {1}'.format(childtb_name,traceback.format_exc()) )
            return False
        else:
            try :
                #print('sql_1 :'+sql_list[1])
                effect_row=self.cursor.execute(sql_list[1])
                self.conn.commit()
            except Exception as e:
                self.conn.rollback()
                logger.info("error sql massage :"+sql_list[1])
                logger.error('delete record failure,\n error massge: {1}'.format(childtb_name,traceback.format_exc()) )
                return False
            else:
                if  effect_row<=0:
                    try :
                        #print('sql_2 :'+sql_list[2])
                        effect_row=self.cursor.execute(sql_list[2])
                        self.conn.commit()
                    except Exception as e:
                        self.conn.rollback()
                        logger.info("error sql massage :"+sql_list[2])
                        logger.error('insert record failure,\n error massge: {1}'.format(childtb_name,traceback.format_exc()) )
                        return False
                    else:
                        return True
                else:
                    return True
    def save_to_mysql(self,item):
        #更新时间
        if  self.conn==None:
            self.conn=init_mysql('119.23.34.166',3306,'pythonspider','python@fly','exchange_rate','utf8')
        #self.cursor=self.conn.cursor()
        if  self.conn !=None:
            self.cursor = self.conn.cursor(cursor=pymysql.cursors.DictCursor)
        self.update_datatime(item['new_update_date'],item['currency_name'])
        #保存数据
        
        size=len(item['data_list'][0])
        #print(size)
        for index in range(0,size):
            ex_cur_name=item['data_list'][0][index][0]
            childtb_name=item['currency_name']+'_'+ex_cur_name+'_tb'
            cur_tb_name=item['currency_name']+'_tb'

            sql_list=[]
            sql='' 
            if  item['currency_name']=='CNY':
                sql="create table if not exists {0}(BuyingRate float(24,2) default null,CashBuyingRate float(24,2) default null, \
                SellingRate float(24,2) default null,CashSellingRate float(24,2) default null,MiddleRate float(24,2) default null, \
                PubTime datetime NOT NULL PRIMARY KEY)".format(childtb_name)
                sql_list.append(sql)
            else:
                sql="create table if not exists {0}(Exchange_Rate float(24,6) default null,PubTime datetime NOT NULL PRIMARY KEY)".format(childtb_name)
                sql_list.append(sql)
                 
            sql="select exchange_currency_name,child_tb_name from {0} where (exchange_currency_name='{1}' AND child_tb_name='{2}')".format(cur_tb_name,ex_cur_name,childtb_name)
            sql_list.append(sql)
            sql="insert into {0}(exchange_currency_name,child_tb_name) VALUES('{1}','{2}')".format(cur_tb_name,ex_cur_name,childtb_name)
            sql_list.append(sql)

            insert_pre_res=self.insert_data_pre(sql_list,childtb_name)
            if  not insert_pre_res:
                continue
            #print(item['data_list'][::-1])

            for i in item['data_list'][::-1]:
                #print(i)
                #print(index)
                line=i[index]
                #print(line)
                if  item['currency_name']=='CNY':
                    sql=" insert into {0}(BuyingRate,CashBuyingRate,SellingRate, \
                    CashSellingRate,MiddleRate,PubTime) VALUES({1},{2},{3},{4},{5},'{6}')".format(childtb_name,line[1],line[2],line[3],line[4],line[5],line[-1])
                else:       
                    sql="insert into {0}(Exchange_Rate,PubTime) VALUES({1},'{2}')".format(childtb_name,line[1],line[2])
                try :
                    #print(sql)
                    effect_row=self.cursor.execute(sql)
                    self.conn.commit()
                except Exception as e:
                    self.conn.rollback()
                    logger.info("error sql massage :"+sql)
                    logger.error('insert record failure,\n error massge: {1}'.format(childtb_name,traceback.format_exc()) )
        if  self.cursor!=None:
            self.cursor.close()
        if  self.conn!=None:
            self.conn.close()
        logger.info('close mysql connect')#日志
    def process_item(self,item,spider):
        self.save_to_excel(item)
        self.save_to_mysql(item)
        
    def close_spider(self,spider):
        logger.info('data save finish')

        
        

    